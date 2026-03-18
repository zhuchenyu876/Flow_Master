"""
为每个意图创建独立的知识库 (One Intent = One Knowledge Base)

核心功能：
- 从 output/step1_processed/intents_*.json 读取意图数据
- 为每个 intent 创建独立的知识库
- 知识库名称：使用 displayName + 语言后缀 + 输入文件名（最长 128 字符，超过自动截断）
- 知识库描述：Intent: {displayName} | ID: {intent_id} | Language: {lang_code}（最长 128 字符，超过自动截断）
- Q&A 数据：
  * Question: intent 的 trainingPhrases（训练短语）
  * Answer: intent 的 displayName（意图名称）

使用方法：
1. 在 .env 文件中配置 API 凭证（ROBOT_KEY, ROBOT_TOKEN, USERNAME）
2. 配置处理参数：
   - CREATE_KB: 是否创建知识库
   - MAX_INTENTS_TO_PROCESS: 限制处理数量（测试时可设为10）
   - KB_NAME_PREFIX / KB_NAME_SUFFIX: 知识库名称前缀/后缀（避免重名）
3. 运行：python create_kb_per_intent.py
4. 选择语言（en / zh / zh-hant / 全部）

输出结果：
- Excel 文件：output/qa_knowledge_bases/temp/qa_{intent_name}_{lang}.xlsx
- 结果 JSON：output/qa_knowledge_bases/kb_per_intent_results_{lang}.json

详细说明：请参阅 KB_PER_INTENT_README.md

作者：chenyu.zhu
日期：2025-12-17
"""

import json
import os
import time
import threading
import hashlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from create_knowledge_base import KnowledgeBaseAPI

from logger_config import get_logger
logger = get_logger(__name__)

# 从 .env 文件读取配置
from dotenv import load_dotenv
load_dotenv()

# ========================================
# 配置信息
# ========================================
# API 配置（从 .env 文件读取，如果没有则使用空字符串）
ROBOT_KEY = os.getenv("ROBOT_KEY", "")
ROBOT_TOKEN = os.getenv("ROBOT_TOKEN", "")
USERNAME = os.getenv("USERNAME", "")

# 是否为每个意图创建知识库（如果 False，只统计）
CREATE_KB = True

# 限制处理的意图数量（None = 全部，或设置具体数字如 10）
MAX_INTENTS_TO_PROCESS = None  # 建议先测试少量意图

# 知识库名称前缀/后缀（用于避免重名冲突）
KB_NAME_PREFIX = "20251212_"  # 添加时间戳前缀避免重名
KB_NAME_SUFFIX = ""  # 已修复embedding参数，不需要后缀了

# 名称与描述长度限制（可通过环境变量覆盖）
MAX_KB_NAME_LENGTH = int(os.getenv("STEP_3_MAX_KB_NAME_LENGTH", "128"))  # 实际限制更严格
MAX_KB_DESC_LENGTH = 128  # API 实际限制


def _sanitize_suffix_component(value: str) -> str:
    """将名称片段转换为仅包含字母、数字、下划线和连字符的安全字符串"""
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in value or "")


def build_kb_name_with_hash(display_name: str, lang_code: str, source_file_tag: str = ""):
    """
    构造知识库名称，包含：
    - 前缀/后缀
    - 语言后缀
    - 输入文件名后缀（防重名）
    - 如果需要截断，添加 hash 保证唯一性

    同时保证总长度不超过 MAX_KB_NAME_LENGTH。

    Args:
        display_name: 意图的 displayName
        lang_code: 语言代码（如 "en", "zh"）
        source_file_tag: 源文件标签（用于避免重名）

    Returns:
        (kb_name, name_truncated)

    示例：
        输入：display_name="VeryLongIntentName...", lang_code="en", source_file_tag="FileA"
        输出：("VeryLongIntentName_en_FileA_a1b2c3d4", True)
                                       ^^^^^^^^ 8位hash保证唯一性
    """
    lang_suffix = f"_{lang_code}" if lang_code else ""
    safe_file_tag = _sanitize_suffix_component(source_file_tag)
    file_suffix = f"_{safe_file_tag}" if safe_file_tag else ""

    # 构造完整的名称（未截断）
    full_name = f"{KB_NAME_PREFIX}{display_name}{KB_NAME_SUFFIX}{lang_suffix}{file_suffix}"

    # 如果不需要截断，直接返回
    if len(full_name) <= MAX_KB_NAME_LENGTH:
        return full_name, False

    # 需要截断，添加 hash 保证唯一性
    # Hash 基于完整名称（包括所有组成部分）
    full_name_for_hash = f"{display_name}|{lang_code}|{source_file_tag}"
    name_hash = hashlib.md5(full_name_for_hash.encode('utf-8')).hexdigest()[:8]

    # Hash 部分：_xxxxxxxx (9个字符)
    HASH_SUFFIX = f"_{name_hash}"
    HASH_LENGTH = len(HASH_SUFFIX)

    # 固定部分长度
    fixed_length = len(KB_NAME_PREFIX) + len(KB_NAME_SUFFIX) + len(lang_suffix) + HASH_LENGTH

    # 剩余可用长度
    remaining = MAX_KB_NAME_LENGTH - fixed_length

    if remaining < 20:
        # 空间不够，优先保证 display_name 和 hash
        # 牺牲 file_suffix
        remaining = MAX_KB_NAME_LENGTH - len(KB_NAME_PREFIX) - len(KB_NAME_SUFFIX) - len(lang_suffix) - HASH_LENGTH
        file_suffix = ""

        if remaining < 10:
            # 还不够，说明前缀后缀太长了
            # 只保留 display_name 的前几个字符 + hash
            remaining = MAX_KB_NAME_LENGTH - HASH_LENGTH - len(lang_suffix)
            trimmed_display = display_name[:max(10, remaining)]
            kb_name = f"{trimmed_display}{lang_suffix}{HASH_SUFFIX}"
            return kb_name[:MAX_KB_NAME_LENGTH], True

    # 为 display_name 和 file_suffix 分配空间
    # 优先保证 display_name，file_suffix 可以被截断或移除
    if len(file_suffix) > remaining // 3:
        # file_suffix 太长，截断或移除
        max_file_suffix_len = remaining // 4
        if max_file_suffix_len < 10:
            file_suffix = ""  # 移除
        else:
            file_suffix = file_suffix[:max_file_suffix_len]

    # 重新计算剩余空间
    remaining = MAX_KB_NAME_LENGTH - len(KB_NAME_PREFIX) - len(KB_NAME_SUFFIX) - len(lang_suffix) - len(file_suffix) - HASH_LENGTH

    # 截断 display_name
    trimmed_display = display_name[:remaining]

    # 构造最终名称
    kb_name = f"{KB_NAME_PREFIX}{trimmed_display}{KB_NAME_SUFFIX}{lang_suffix}{file_suffix}{HASH_SUFFIX}"

    # 最终保险：确保不超出上限
    if len(kb_name) > MAX_KB_NAME_LENGTH:
        kb_name = kb_name[:MAX_KB_NAME_LENGTH]

    return kb_name, True


def build_kb_name(display_name: str, lang_code: str, source_file_tag: str = ""):
    """
    构造知识库名称，使用改进的 hash 算法保证唯一性

    Args:
        display_name: 意图的 displayName
        lang_code: 语言代码
        source_file_tag: 源文件标签

    Returns:
        (kb_name, name_truncated)
    """
    # 使用改进的命名函数（带 hash 保证唯一性）
    return build_kb_name_with_hash(display_name, lang_code, source_file_tag)

# 语言代码
LANGUAGE_CODE = "en"  # 可选：en, zh, zh-hant

# 并发上传知识库的线程数（默认一次 3 个）
UPLOAD_WORKERS = 3  # 固定为3个并发线程


def create_kb_for_single_intent(api: KnowledgeBaseAPI, intent: dict, lang_code: str, source_file_tag: str = ""):
    """
    为单个意图创建知识库并上传 Q&A

    Args:
        api: API 客户端
        intent: 意图数据
        lang_code: 语言代码
        source_file_tag: 输入文件名标签（用于防重名，会添加到知识库名称末尾）

    Returns:
        元组 (kb_id, kb_name_used, name_truncated, is_reused, error_msg)：
        知识库ID、实际使用的知识库名称、是否截断、是否复用已存在的知识库、错误信息（成功时为None）
        或 (None, None, name_truncated, False, error_msg)（失败/跳过）
    """
    intent_id = intent.get('id', '')
    display_name = intent.get('displayName', '')
    training_phrases = intent.get('trainingPhrases', [])
    name_truncated = False
    
    if not display_name:
        error_msg = f"意图名称为空"
        logger.debug(f"跳过：{error_msg}")
        return None, None, name_truncated, False, error_msg

    # 如果没有训练短语，仍然创建知识库（但不上传Q&A）
    has_training_phrases = bool(training_phrases)
    if not has_training_phrases:
        logger.info(f"   ℹ️  意图 {display_name} 无训练短语，将创建空知识库（不上传Q&A）")
    
    # 构造知识库名称（含语言、文件后缀）并控制长度
    kb_name, name_truncated = build_kb_name(display_name, lang_code, source_file_tag)
    logger.debug(f"   📝 知识库名称构建: {kb_name}")
    logger.debug(f"      - 意图名: {display_name}")
    logger.debug(f"      - 语言: {lang_code}")
    logger.debug(f"      - Agent/文件标签: {source_file_tag}")
    logger.debug(f"      - 是否截断: {name_truncated}")
    if name_truncated:
        logger.warning(f"   ⚠️  名称过长，已截断为 {MAX_KB_NAME_LENGTH} 字符")
        logger.warning(f"      原始意图名: {display_name}")
        logger.warning(f"      文件后缀: {_sanitize_suffix_component(source_file_tag)}")
        logger.warning(f"      最终名称: {kb_name}")
    
    # 知识库描述：包含 intent ID 和完整名称（描述长度限制128字符）
    kb_description = f"Intent: {display_name} | ID: {intent_id} | Language: {lang_code}"
    
    # 如果描述过长，截断（这是关键！）
    if len(kb_description) > MAX_KB_DESC_LENGTH:
        kb_description = kb_description[:MAX_KB_DESC_LENGTH]
        print(f"   ⚠️  描述过长（{len(kb_description)} → {MAX_KB_DESC_LENGTH} 字符），已截断")
    
    # print(f"\n{'='*80}")
    # print(f"📍 意图: {display_name}")
    # print(f"   Intent ID: {intent_id}")
    # print(f"   训练短语数量: {len(training_phrases)}")
    # print(f"{'='*80}")
    
    # 根据语言代码选择合适的 embedding 参数
    # 英文和中文都使用 bge-m3 模型（多语言模型）
    # 但 emb_language 要与内容语言匹配
    if lang_code == "en":
        emb_language = "en"
        emb_model = "bge-m3"  # 模型名称用中文的，但支持英文
    elif lang_code in ["zh-hant"]:
        # zh-hant (粤语/繁体中文) 使用独立的 'zh-hant'
        emb_language = "zh-hant"
        emb_model = "bge-m3"
    elif lang_code == "zh":
        emb_language = "zh"
        emb_model = "bge-m3"
    else:
        # 默认英文
        emb_language = "en"
        emb_model = "bge-m3"
    
    # 创建知识库
    logger.debug(f"   正在创建知识库: {kb_name}")
    logger.debug(f"   描述: {kb_description}")
    logger.debug(f"   Embedding: {emb_language} / {emb_model}")
    
    result = api.create_knowledge_base(
        name=kb_name,
        description=kb_description,
        emb_language=emb_language,
        emb_model=emb_model
    )

    # 检查 result 是否为字典类型
    if not isinstance(result, dict):
        error_msg = f"API 返回格式错误：期望 dict，实际为 {type(result).__name__}"
        if isinstance(result, str):
            error_msg += f"，内容: {result[:200]}"
        logger.error(f"创建失败: {error_msg}")
        return None, None, name_truncated, False, error_msg

    # 确保 result 是字典类型
    if not isinstance(result, dict):
        error_msg = f"API 返回格式错误：期望 dict，实际为 {type(result).__name__}"
        if isinstance(result, str):
            error_msg += f"，内容: {result[:200]}"
        logger.error(f"创建失败: {error_msg}")
        return None, None, name_truncated, False, error_msg

    result_code = result.get("code")
    if result_code == "000000":
        kb_id = str(result.get("data", {}).get("id"))
        logger.info(f"✅ 知识库创建成功: {kb_name} (ID: {kb_id})")
    elif result_code == "EXISTS":
        # 知识库已存在，复用已存在的ID，跳过Q&A上传
        kb_id = str(result.get("existing_kb_id"))
        kb_name = kb_name  # 保持原名称
        logger.info(f"✅ 知识库已存在，复用: {kb_name} (ID: {kb_id})")
        print(f"   ℹ️  跳过 Q&A 上传（复用现有知识库）")
        return kb_id, kb_name, name_truncated, True, None
    elif result_code == "400001":
        # 兼容旧的400001错误码，复用已存在的ID，跳过Q&A上传
        data = result.get("data", {})
        kb_id = None
        
        if isinstance(data, dict):
            # data 是字典，直接获取 ID
            kb_id = data.get("id")
            if kb_id:
                kb_id = str(kb_id)
        elif isinstance(data, str):
            # data 是字符串，可能是错误信息，尝试通过名称查询
            logger.debug(f"   ℹ️  400001 错误，data 是字符串: {data[:100]}")
            logger.info(f"   🔍 检测到重复知识库，正在查询已存在的知识库ID...")
            existing_kb = api.check_kb_exists_by_name(kb_name)
            if existing_kb:
                kb_id = str(existing_kb.get("id"))
                logger.info(f"   ✅ 找到已存在的知识库: {kb_name} (ID: {kb_id})")
            else:
                logger.warning(f"   ⚠️  无法通过名称找到已存在的知识库")
        
        if not kb_id:
            # 如果仍然无法获取 ID，尝试通过名称查询
            logger.info(f"   🔍 尝试通过名称查询已存在的知识库ID...")
            existing_kb = api.check_kb_exists_by_name(kb_name)
            if existing_kb:
                kb_id = str(existing_kb.get("id"))
                logger.info(f"   ✅ 找到已存在的知识库: {kb_name} (ID: {kb_id})")
            else:
                error_msg = f"400001 错误但无法获取知识库ID，data格式: {type(data).__name__}"
                if isinstance(data, str):
                    error_msg += f"，内容: {data[:100]}"
                logger.error(f"创建失败: {error_msg}")
                return None, None, name_truncated, False, error_msg
        
        kb_name = kb_name  # 保持原名称
        logger.info(f"✅ 知识库已存在，复用: {kb_name} (ID: {kb_id})")
        print(f"   ℹ️  跳过 Q&A 上传（复用现有知识库）")
        return kb_id, kb_name, name_truncated, True, None
    else:
        # 确保 result 是字典类型
        if isinstance(result, dict):
            error_msg = result.get('message', '未知错误')
            error_code = result.get('code', 'UNKNOWN')
        else:
            error_msg = f"API返回格式错误：{type(result).__name__}"
            if isinstance(result, str):
                error_msg += f"，内容: {result[:200]}"
            error_code = 'FORMAT_ERROR'
        full_error_msg = f"API返回错误 (code: {error_code}): {error_msg}"
        logger.error(f"创建失败: {full_error_msg}")
        # 检查是否是"名称已存在"错误（兼容旧版本）
        if "already exists" in error_msg.lower() or "already exist" in error_msg.lower():
            # 尝试查询已存在的知识库ID
            logger.info(f"   🔍 检测到重复知识库，正在查询已存在的知识库ID...")
            existing_kb = api.check_kb_exists_by_name(kb_name)
            if existing_kb:
                kb_id = str(existing_kb.get("id"))
                logger.info(f"   ✅ 找到已存在的知识库: {kb_name} (ID: {kb_id})")
                logger.info(f"   ℹ️  跳过 Q&A 上传（复用现有知识库）")
                return kb_id, kb_name, name_truncated, True, None
            else:
                logger.warning(f"   ⚠️  无法找到已存在的知识库，返回DUPLICATE标记")
                return None, "DUPLICATE", name_truncated, False, full_error_msg
        return None, None, name_truncated, False, full_error_msg
    
    # 准备 Q&A 数据并生成 Excel
    import pandas as pd
    from openpyxl.styles import Font, PatternFill
    
    qa_pairs = []
    for phrase in training_phrases:
        if phrase and phrase.strip():
            qa_pairs.append({
                "question": phrase.strip(),
                "answer": display_name
            })
    
    if not qa_pairs:
        # 如果没有有效的训练短语，但知识库已经创建成功，则不上传Q&A
        logger.info(f"   ℹ️  知识库已创建（ID: {kb_id}），但无有效训练短语，跳过Q&A上传")
        return kb_id, kb_name, name_truncated, False, None
    
    # 生成临时 Excel 文件（文件名使用 displayName）
    temp_excel_dir = os.getenv("STEP_3_QA_TEMP_DIR", "output/qa_knowledge_bases/temp")

    # 确保temp_excel_dir是绝对路径
    if not os.path.isabs(temp_excel_dir):
        # 如果是相对路径，相对于项目根目录（step3_kb_creator.py 所在目录）
        # write by senlin.deng 2025-12-23
        project_root = os.path.dirname(os.path.abspath(__file__))
        temp_excel_dir = os.path.join(project_root, temp_excel_dir)

    # 确保目录存在，使用更安全的目录创建方式
    try:
        os.makedirs(temp_excel_dir, exist_ok=True)
        logger.debug(f"临时目录创建成功: {temp_excel_dir}")
    except OSError as e:
        logger.warning(f"创建临时目录失败，使用系统临时目录: {e}")
        import tempfile
        temp_excel_dir = tempfile.gettempdir()
        os.makedirs(temp_excel_dir, exist_ok=True)

    # 清理文件名中的特殊字符，并限制长度避免Windows路径过长问题
    # 首先替换特殊字符为空格，然后清理所有非字母数字字符
    cleaned_name = display_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace(' ', '_')
    safe_display_name = "".join(c for c in cleaned_name if c.isalnum() or c == '_')

    # 如果清理后为空，使用默认名称
    if not safe_display_name:
        safe_display_name = "intent"

    # 限制长度到20个字符（极度保守的限制）
    if len(safe_display_name) > 20:
        safe_display_name = safe_display_name[:17] + "..."

    logger.debug(f"原始名称: {display_name}")
    logger.debug(f"清理后名称: {safe_display_name}")

    # 生成唯一的临时文件名，避免并发冲突
    import uuid
    unique_id = str(uuid.uuid4())[:8]
    temp_excel_file = os.path.join(temp_excel_dir, f"qa_{safe_display_name}_{lang_code}_{unique_id}.xlsx")

    # 最终确保目录存在
    try:
        os.makedirs(os.path.dirname(temp_excel_file), exist_ok=True)
    except OSError as e:
        logger.error(f"无法创建文件目录: {e}")
        raise
    
    df = pd.DataFrame(qa_pairs)
    
    try:
        writer = pd.ExcelWriter(temp_excel_file, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Q&A')
        
        # 应用样式
        workbook = writer.book
        sheet = writer.sheets['Q&A']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col_num, column in enumerate(sheet.columns, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            sheet.column_dimensions[chr(64 + col_num)].width = 50
        
        writer.close()
        
        print(f"   ✅ Excel 已生成: {os.path.basename(temp_excel_file)}")
        
        # 上传 Excel 到知识库
        logger.info(f"   ⬆️  正在上传 {len(qa_pairs)} 条 Q&A 到知识库 {kb_id}...")
        
        upload_result = api.import_qa_from_file(
            username=USERNAME,
            knowledge_base_id=kb_id,
            folder_id="root",
            file_path=temp_excel_file
        )
        
        if upload_result.get("code") == "000000":
            logger.info(f"   ✅ Q&A 上传成功！")
            logger.info(f"   📁 Excel 文件已保存: {temp_excel_file}")
            # Excel 文件保留在临时目录中，不删除，方便后续查看和调试
            return kb_id, kb_name, name_truncated, False, None
        else:
            error_msg = f"Q&A 上传失败: {upload_result.get('message', '未知错误')}"
            logger.error(f"   ❌ {error_msg}")
            return kb_id, kb_name, name_truncated, False, error_msg
            
    except Exception as e:
        import traceback
        error_msg = f"Excel 生成或上传失败: {str(e)}\n{traceback.format_exc()}"
        logger.error(f"   ❌ {error_msg}")
        return kb_id, kb_name, name_truncated, False, error_msg

    return kb_id, kb_name, name_truncated, False, None


def process_language_file(intents_file: str, lang_code: str, task_id: str = None, db_session = None, output_dir: str = None):
    """
    处理单个语言的 intents 文件
    
    Args:
        intents_file: intents JSON 文件路径
        lang_code: 语言代码
        task_id: 任务ID（可选，用于写入数据库）
        db_session: 数据库会话（可选）
        output_dir: 输出目录（可选，默认使用环境变量）
    """
    
    # 检查必要的依赖
    try:
        import pandas as pd
        import openpyxl
    except ImportError as e:
        logger.error(f"❌ 缺少必要的依赖库: {e}")
        logger.error("   请运行: pip install pandas openpyxl")
        return
    
    # 检查 API 配置
    if CREATE_KB:
        if not ROBOT_KEY or not ROBOT_TOKEN or not USERNAME:
            logger.error("❌ API 配置缺失！")
            logger.error("   请在 .env 文件中设置 ROBOT_KEY, ROBOT_TOKEN, USERNAME")
            return
    
    # 检查文件是否存在
    if not os.path.exists(intents_file):
        print(f"❌ 文件不存在: {intents_file}")
        return

    # 优先使用环境变量中的原始 exported_flow 文件名，如果没有则使用 intents 文件名
    source_file_tag = os.getenv('STEP_3_SOURCE_FILE_TAG', '')
    if not source_file_tag:
        # 如果没有设置环境变量，使用 intents 文件名（向后兼容）
        source_file_tag = os.path.splitext(os.path.basename(intents_file))[0]
        logger.info(f"⚠️  未设置 STEP_3_SOURCE_FILE_TAG，使用 intents 文件名: {source_file_tag}")
    else:
        logger.info(f"📋 使用原始 exported_flow 文件名作为知识库后缀: {source_file_tag}")
    
    # 读取 intents 数据
    try:
        with open(intents_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"❌ 读取文件失败: {str(e)}")
        return
    
    # 支持两种格式：{"intents": [...]} 或直接 [...]
    if isinstance(data, dict) and 'intents' in data:
        intents = data['intents']
    elif isinstance(data, list):
        intents = data
    else:
        print(f"❌ 无效的 intents 数据格式")
        return
    
    if not intents:
        print(f"⚠️  未找到 intents 数据")
        return
    
    # 过滤掉非字典类型的元素（防止数据格式错误）
    original_count = len(intents)
    intents = [intent for intent in intents if isinstance(intent, dict)]
    if len(intents) < original_count:
        skipped_count = original_count - len(intents)
        logger.warning(f"⚠️  跳过了 {skipped_count} 个非字典类型的 intent 数据")
    
    total_intents = len(intents)
    logger.info(f"📊 找到 {total_intents} 个意图")
    
    # 应用限制
    if MAX_INTENTS_TO_PROCESS and MAX_INTENTS_TO_PROCESS < total_intents:
        print(f"⚠️  限制处理数量: 仅处理前 {MAX_INTENTS_TO_PROCESS} 个意图")
        print(f"   （修改脚本中的 MAX_INTENTS_TO_PROCESS 变量可以处理更多）")
        intents = intents[:MAX_INTENTS_TO_PROCESS]
    
    # 创建 API 客户端
    api = KnowledgeBaseAPI(ROBOT_KEY, ROBOT_TOKEN, USERNAME)
    
    # **NEW: 批量检查空间中是否已有重复知识库（检查所有intent）**
    # 一次性检查所有待创建的知识库名称是否已存在
    existing_kb_names_in_space = set()  # 存储空间中所有知识库名称
    if CREATE_KB:
        logger.info("🔍 预检查：验证空间中是否已有重复知识库...")
        
        # 生成所有待创建的知识库名称（包含语言、文件后缀）
        all_kb_names_to_create = []
        for intent in intents:
            display_name = intent.get('displayName', '')
            if display_name:
                kb_name, _ = build_kb_name(display_name, lang_code, source_file_tag)
                all_kb_names_to_create.append((display_name, kb_name))
        
        logger.info(f"   待检查: {len(all_kb_names_to_create)} 个 intent 的知识库名称")
        logger.info(f"   示例名称: {all_kb_names_to_create[0][1] if all_kb_names_to_create else 'N/A'}...")
        
        # 一次性获取空间中的所有知识库
        kb_list_result = api.list_knowledge_bases(page=1, size=200)
        if kb_list_result.get("code") == "000000":
            kb_list_data = kb_list_result.get("data", {})
            existing_kbs = kb_list_data.get("list", [])
            total_kbs = kb_list_data.get("total", 0)
            
            existing_kb_names_in_space = {kb.get("knowledge_base_name") for kb in existing_kbs}
            
            # 如果总数超过200，继续查询（支持分页）
            if total_kbs > 200:
                for page in range(2, (total_kbs // 200) + 2):
                    result = api.list_knowledge_bases(page=page, size=200)
                    if result.get("code") == "000000":
                        more_kbs = result.get("data", {}).get("list", [])
                        existing_kb_names_in_space.update({kb.get("knowledge_base_name") for kb in more_kbs})
            
            logger.info(f"   已加载 {len(existing_kb_names_in_space)} 个知识库名称")
            
            # 检查所有待创建的知识库名称中有多少已存在
            duplicates_found = [(display_name, kb_name) for display_name, kb_name in all_kb_names_to_create 
                               if kb_name in existing_kb_names_in_space]
            
            if duplicates_found:
                for display_name, kb_name in duplicates_found[:10]:
                    logger.warning(f"  - {display_name} → {kb_name}")
                if len(duplicates_found) > 10:
                    logger.warning(f"  ... 还有 {len(duplicates_found)-10} 个重复")
                logger.warning(f"\n可能原因:")
                logger.warning(f"  1. 之前的任务已上传，但数据库映射记录丢失")
                logger.warning(f"  2. 手动创建了同名知识库")
                logger.warning(f"  3. 数据库被清空但空间中的知识库未删除")
                logger.warning(f"\n建议操作:")
                logger.warning(f"  1. 在赛博坦后台删除重复的知识库，然后重新运行")
                logger.warning(f"  2. 或在数据库中手动创建映射记录（需要知道所有 KB ID）")
                logger.warning(f"  3. 或修改 KB_NAME_PREFIX/SUFFIX 以避免命名冲突")
                logger.warning(f"\n⏭️  为避免重复创建错误，已跳过本批次所有知识库上传。")
                logger.warning("="*80 + "\n")
                print("\n" + "="*80)
                print(f"⚠️  知识库上传已跳过（检测到 {len(duplicates_found)} 个重复）")
                print("="*80 + "\n")
                return  # 直接返回，不创建任何知识库
            else:
                logger.info(f"   ✅ 预检查通过，所有 {len(all_kb_names_to_create)} 个知识库名称均未重复")
        else:
            logger.warning(f"⚠️  无法获取空间知识库列表: {kb_list_result.get('message')}")
            logger.warning(f"   将在创建过程中检测重复，一旦发现立即停止整个批次")
    
    # 输出目录配置（优先使用传入参数，其次环境变量，最后默认值）
    if output_dir is None:
        output_dir = os.getenv("STEP_3_QA_OUTPUT_DIR", "output/qa_knowledge_bases")
    output_file = os.path.join(output_dir, f"kb_per_intent_results_{lang_code}.json")
    qa_temp_dir = os.getenv("STEP_3_QA_TEMP_DIR", os.path.join(output_dir, "temp"))
    
    # ========================================
    # 🔥 新增逻辑：复用已有映射，避免重复创建
    # ========================================
    existing_results = {}
    reused_from_file = 0
    
    # 检查是否已有映射文件（由 step3_kb_reuse_helper 生成）
    if os.path.exists(output_file):
        try:
            with open(output_file, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
            existing_results = existing_data.get('results', {})
            reused_from_file = len(existing_results)
            
            if existing_results:
                logger.info(f"📂 检测到已有映射文件: {output_file}")
                logger.info(f"   已有映射数量: {reused_from_file}")
                
                # 筛选出还没有映射的 intents
                intents_to_process = []
                for intent in intents:
                    display_name = intent.get('displayName', '')
                    if display_name and display_name not in existing_results:
                        intents_to_process.append(intent)
                
                if not intents_to_process:
                    logger.info(f"   ✅ 所有 {total_intents} 个 intents 都已有映射，跳过创建步骤")
                    logger.info(f"   💾 映射文件: {output_file}")
                    return
                else:
                    logger.info(f"   🔄 需要处理的新 intents: {len(intents_to_process)}/{total_intents}")
                    intents = intents_to_process  # 只处理没有映射的 intents
        except Exception as e:
            logger.warning(f"   ⚠️  读取已有映射文件失败: {e}")
            existing_results = {}
    
    if not existing_results:
        logger.info("   ⚠️  注意：遇到第一个重复知识库将立即停止整个批次")
    
    # 处理每个意图（支持并发上传）
    logger.info("")
    logger.info("="*80)
    logger.info(f"🚀 开始处理 {len(intents)} 个新 intents")
    logger.info("="*80)
    results = existing_results.copy()  # 从已有结果开始
    success_count = 0
    failed_count = 0
    skipped_count = 0
    duplicate_stop_event = threading.Event()
    

    def process_single_intent(idx, intent):
        """在工作线程中处理单个 intent，返回结果字典"""
        # 检查 intent 是否为字典类型
        if not isinstance(intent, dict):
            error_msg = f"intent 数据格式错误：期望 dict，实际为 {type(intent).__name__}"
            logger.error(f"   ❌ [{idx}] {error_msg}")
            if isinstance(intent, str):
                logger.error(f"      实际值: {intent[:100]}...")  # 只显示前100个字符
            return {
                "display_name": f"intent_{idx}",
                "intent": intent if isinstance(intent, dict) else {},
                "status": "error",
                "kb_id": None,
                "kb_name": None,
                "error": error_msg
            }
        
        display_name = intent.get('displayName', f'intent_{idx}')
        
        # 初始化返回变量，防止异常时未定义
        kb_id = None
        kb_name_used = None
        name_truncated = False
        is_reused = False
        
        # 如果已有重复标记，跳过后续任务
        if duplicate_stop_event.is_set():
            return {
                "display_name": display_name,
                "intent": intent,
                "status": "skipped_due_to_duplicate",
                "kb_id": None,
                "kb_name": None
            }
        
        # 逐个检查是否已存在同名知识库（结合预检查结果）
        lang_suffix_check = f"_{lang_code}" if lang_code else ""
        if CREATE_KB and existing_kb_names_in_space:
            kb_name_to_check, _ = build_kb_name(display_name, lang_code, source_file_tag)
            if kb_name_to_check in existing_kb_names_in_space:
                duplicate_stop_event.set()
                return {
                    "display_name": display_name,
                    "intent": intent,
                    "status": "duplicate_detected",
                    "kb_id": None,
                    "kb_name": kb_name_to_check
                }
        
        if not CREATE_KB:
            return {
                "display_name": display_name,
                "intent": intent,
                "status": "stats_only",
                "kb_id": None,
                "kb_name": None
            }
        
        try:
            # 独立的 API 客户端，避免跨线程共享状态
            api_client = KnowledgeBaseAPI(ROBOT_KEY, ROBOT_TOKEN, USERNAME)
            logger.info(f"   [线程] 开始处理 intent: {display_name}")
            kb_id, kb_name_used, name_truncated, is_reused, error_msg = create_kb_for_single_intent(
                api_client, intent, lang_code, source_file_tag
            )
            
            # 检测重复错误：立即通知其他线程停止
            if kb_name_used == "DUPLICATE":
                duplicate_stop_event.set()
                return {
                    "display_name": display_name,
                    "intent": intent,
                    "status": "duplicate_detected",
                    "kb_id": None,
                    "kb_name": kb_name_used,
                    "error": "知识库名称已存在，但无法获取已存在的知识库ID"
                }

            if kb_id:
                # 知识库创建成功（或复用成功）
                if is_reused:
                    logger.info(f"   [线程] intent {display_name} 复用已存在的知识库 (ID: {kb_id})")
                else:
                    logger.info(f"   [线程] intent {display_name} 知识库创建成功 (ID: {kb_id})")

                # 检查是否有训练短语
                has_phrases = bool(intent.get('trainingPhrases'))
                if not has_phrases:
                    logger.info(f"   [线程] intent {display_name} 无训练短语，知识库创建完成（不上传Q&A）")
                elif error_msg:
                    logger.warning(f"   [线程] intent {display_name} 知识库已创建，但Q&A上传失败: {error_msg}")
                    return {
                        "display_name": display_name,
                        "intent": intent,
                        "status": "failed",
                        "kb_id": kb_id,  # 知识库已创建
                        "kb_name": kb_name_used,
                        "name_truncated": name_truncated,
                        "error": f"Q&A上传失败: {error_msg}",
                        "kb_created": True  # 标记知识库已创建
                    }

                # 避免请求过快（轻量节流）
                time.sleep(0.3)

                return {
                    "display_name": display_name,
                    "intent": intent,
                    "status": "success",
                    "kb_id": kb_id,
                    "kb_name": kb_name_used,
                    "name_truncated": name_truncated,
                    "reused": is_reused
                }

            # 知识库创建失败
            error_info = error_msg or "知识库创建失败，原因未知"
            logger.error(f"   [线程] intent {display_name} 知识库创建失败: {error_info}")
            return {
                "display_name": display_name,
                "intent": intent,
                "status": "failed",
                "kb_id": None,
                "kb_name": kb_name_used,
                "name_truncated": name_truncated,
                "error": f"知识库创建失败: {error_info}",
                "kb_created": False  # 标记知识库未创建
            }
        except Exception as e:
            import traceback
            error_detail = f"{str(e)}\n{traceback.format_exc()}"
            logger.error(f"   [线程] 处理 intent {display_name} 时发生异常: {error_detail}")
            return {
                "display_name": display_name,
                "intent": intent,
                "status": "error",
                "kb_id": None,
                "kb_name": None,
                "name_truncated": name_truncated,
                "error": str(e),
                "error_detail": error_detail
            }
    
    # 使用并发或串行处理
    if CREATE_KB and UPLOAD_WORKERS > 1 and len(intents) > 1:
        logger.info(f"🚀 开始并发处理 {len(intents)} 个 intents（并发数: {UPLOAD_WORKERS}）")
        processed_count = 0
        with ThreadPoolExecutor(max_workers=UPLOAD_WORKERS) as executor:
            future_map = {executor.submit(process_single_intent, idx, intent): (idx, intent)
                          for idx, intent in enumerate(intents, 1)}
            for future in as_completed(future_map):
                processed_count += 1
                res = future.result()
                display_name = res["display_name"]
                intent = res["intent"]
                status_msg = f"状态: {res['status']}"
                if res.get("error"):
                    status_msg += f" - 错误: {res['error']}"
                logger.info(f"   [{processed_count}/{len(intents)}] 处理完成: {display_name} - {status_msg}")
                
                if res["status"] == "success":
                    success_count += 1
                    is_reused = res.get("reused", False)
                    results[display_name] = {
                        "intent_id": intent.get('id', ''),
                        "display_name": display_name,
                        "kb_id": res["kb_id"],
                        "kb_name": res["kb_name"],
                        "name_truncated": res.get("name_truncated", False),
                        "status": "success",
                        "reused": is_reused,
                        "training_phrases_count": len(intent.get('trainingPhrases', []))
                    }
                elif res["status"] == "stats_only":
                    results[display_name] = {
                        "training_phrases_count": len(intent.get('trainingPhrases', []))
                    }
                elif res["status"] == "duplicate_detected":
                    failed_count += 1
                    results[display_name] = {
                        "intent_id": intent.get('id', ''),
                        "display_name": display_name,
                        "kb_id": None,
                        "status": "duplicate",
                        "training_phrases_count": len(intent.get('trainingPhrases', []))
                    }
                elif res["status"] == "skipped_due_to_duplicate":
                    results[display_name] = {
                        "intent_id": intent.get('id', ''),
                        "display_name": display_name,
                        "kb_id": None,
                        "status": "skipped",
                        "training_phrases_count": len(intent.get('trainingPhrases', []))
                    }
                else:
                    failed_count += 1
                    error_msg = res.get("error", "未知错误")
                    logger.error(f"   ❌ [{processed_count}/{len(intents)}] 处理失败: {display_name}")
                    logger.error(f"      错误信息: {error_msg}")
                    results[display_name] = {
                        "intent_id": intent.get('id', ''),
                        "display_name": display_name,
                        "kb_id": None,
                        "status": res["status"],
                        "error": error_msg,
                        "training_phrases_count": len(intent.get('trainingPhrases', []))
                    }

    else:
        logger.info(f"🚀 开始串行处理 {len(intents)} 个 intents")
        for idx, intent in enumerate(intents, 1):
            display_name = intent.get('displayName', f'intent_{idx}')
            logger.info(f"   [{idx}/{len(intents)}] 正在处理: {display_name}")
            res = process_single_intent(idx, intent)
            display_name = res["display_name"]
            status_msg = f"状态: {res['status']}"
            if res.get("error"):
                status_msg += f" - 错误: {res['error']}"
            logger.info(f"   [{idx}/{len(intents)}] 处理完成: {display_name} - {status_msg}")
            
            if res["status"] == "success":
                success_count += 1
                is_reused = res.get("reused", False)
                results[display_name] = {
                    "intent_id": intent.get('id', ''),
                    "display_name": display_name,
                    "kb_id": res["kb_id"],
                    "kb_name": res["kb_name"],
                    "name_truncated": res.get("name_truncated", False),
                    "status": "success",
                    "reused": is_reused,
                    "training_phrases_count": len(intent.get('trainingPhrases', []))
                }
                # 避免请求过快
                if idx < len(intents):
                    time.sleep(0.3)
            elif res["status"] == "stats_only":
                results[display_name] = {
                    "training_phrases_count": len(intent.get('trainingPhrases', []))
                }
            elif res["status"] == "duplicate_detected":
                failed_count += 1
                results[display_name] = {
                    "intent_id": intent.get('id', ''),
                    "display_name": display_name,
                    "kb_id": None,
                    "status": "duplicate",
                    "training_phrases_count": len(intent.get('trainingPhrases', []))
                }
                print("\n" + "="*80)
                print(f"⚠️  知识库上传已停止（第 {idx} 个 intent 检测到重复）")
                print("="*80 + "\n")
                break
            else:
                failed_count += 1
                error_msg = res.get("error", "未知错误")
                logger.error(f"   ❌ [{idx}/{len(intents)}] 处理失败: {display_name}")
                logger.error(f"      错误信息: {error_msg}")
                results[display_name] = {
                    "intent_id": intent.get('id', ''),
                    "display_name": display_name,
                    "kb_id": None,
                    "status": res["status"],
                    "error": error_msg,
                    "training_phrases_count": len(intent.get('trainingPhrases', []))
                }
                # 避免请求过快
                if idx < len(intents):
                    time.sleep(0.3)

    # 并发模式下，如果检测到重复，标记停止信息
    if duplicate_stop_event.is_set():
        logger.warning("\n" + "="*80)
        logger.warning("⚠️  检测到知识库重复，已停止后续上传")
        logger.warning("="*80 + "\n")
    
    # 将成功结果写入数据库（单线程，避免 session 跨线程问题）
    # 优化：批量提交，避免每个知识库都单独 commit，减少数据库操作阻塞
    if CREATE_KB and task_id and db_session:
        try:
            from database import crud
            from database.models import KBStatus, KnowledgeBaseMapping
        except Exception as e:
            logger.warning(f"⚠️  无法加载数据库模块，跳过写入: {e}")
        else:
            logger.info(f"💾 开始批量写入数据库映射（共 {len(results)} 条记录）...")
            success_records = []
            failed_records = []
            
            for display_name, record in results.items():
                kb_id = record.get("kb_id")
                if record.get("status") != "success" or not kb_id:
                    continue
                success_records.append((display_name, record))
            
            # 批量处理，每批提交一次（减少 commit 次数）
            batch_size = 50  # 每批处理 50 条
            total_batches = (len(success_records) + batch_size - 1) // batch_size
            
            for batch_idx in range(total_batches):
                start_idx = batch_idx * batch_size
                end_idx = min(start_idx + batch_size, len(success_records))
                batch_records = success_records[start_idx:end_idx]
                
                try:
                    for display_name, record in batch_records:
                        kb_id = record.get("kb_id")
                        existing = db_session.query(KnowledgeBaseMapping).filter_by(
                            task_id=task_id,
                            intent_name=display_name,
                            language=lang_code
                        ).first()
                        
                        if existing:
                            crud.update_kb_mapping(
                                db=db_session,
                                task_id=task_id,
                                intent_name=display_name,
                                language=lang_code,
                                status=KBStatus.CREATED,
                                kb_id=kb_id,
                                kb_name=record.get("kb_name")
                            )
                        else:
                            crud.create_kb_mapping(
                                db=db_session,
                                task_id=task_id,
                                intent_name=display_name,
                                language=lang_code,
                                intent_id=record.get("intent_id", ""),
                                qa_count=record.get("training_phrases_count", 0)
                            )
                            crud.update_kb_mapping(
                                db=db_session,
                                task_id=task_id,
                                intent_name=display_name,
                                language=lang_code,
                                status=KBStatus.CREATED,
                                kb_id=kb_id,
                                kb_name=record.get("kb_name")
                            )
                    
                    # 批量提交（每批提交一次，而不是每个知识库都提交）
                    db_session.commit()
                    logger.debug(f"   ✅ 批次 {batch_idx + 1}/{total_batches} 数据库写入成功 ({len(batch_records)} 条)")
                except Exception as e:
                    logger.warning(f"      ⚠️  批次 {batch_idx + 1}/{total_batches} 数据库写入失败: {e}")
                    db_session.rollback()
                    # 记录失败的记录，但不中断流程
                    failed_records.extend(batch_records)
            
            if failed_records:
                logger.warning(f"⚠️  共有 {len(failed_records)} 条记录数据库写入失败，但流程继续")
            else:
                logger.info(f"✅ 所有 {len(success_records)} 条数据库映射已成功写入")
    
    # 显示结果汇总
    logger.info("")
    logger.info("="*80)
    logger.info("📊 结果汇总")
    logger.info("="*80)
    logger.info(f"总意图数: {total_intents}")
    
    if reused_from_file > 0:
        logger.info(f"📂 从映射文件复用: {reused_from_file} 个")
        logger.info(f"🔄 本次需要处理: {len(intents)} 个")
    else:
        logger.info(f"已处理: {len(intents)}")
    
    if CREATE_KB and len(intents) > 0:
        logger.info(f"✅ 本次创建成功: {success_count}/{len(intents)} 个知识库")
        if skipped_count > 0:
            logger.info(f"   ├─ 本次新创建: {success_count - skipped_count} 个")
            logger.info(f"   └─ 跳过（已存在）: {skipped_count} 个")
        logger.info(f"❌ 本次创建失败: {failed_count}/{len(intents)} 个知识库")
    
    # 保存结果到文件
    os.makedirs(output_dir, exist_ok=True)
    
    # 计算最终的失败数量
    final_success_count = sum(1 for r in results.values() if r.get("status") == "success" and r.get("kb_id"))
    final_failed_count = sum(1 for r in results.values() if r.get("status") in ["failed", "error"])
    
    # 保存结果
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            "language": lang_code,
            "total_intents": total_intents,  # 使用原始的总数
            "success_count": final_success_count,
            "failed_count": final_failed_count,
            "reused_count": reused_from_file,  # 新增：记录复用数量
            "results": results
        }, f, indent=2, ensure_ascii=False)
    
    logger.info(f"\n💾 结果已保存到: {output_file}")
    logger.info(f"   总映射数: {len(results)} 个")
    if reused_from_file > 0:
        logger.info(f"   ├─ 从文件复用: {reused_from_file} 个")
        logger.info(f"   └─ 本次创建: {final_success_count - reused_from_file} 个")
    else:
        logger.info(f"   成功创建: {final_success_count} 个知识库")
    if final_failed_count > 0:
        logger.info(f"   创建失败: {final_failed_count} 个知识库")
    logger.info("="*80)


def main():
    """主函数"""
    
    # 选择要处理的语言
    languages = [
        {"code": "en", "file": "output/step1_processed/intents_en.json", "name": "English"},
        {"code": "zh", "file": "output/step1_processed/intents_zh.json", "name": "简体中文"},
        {"code": "zh-hant", "file": "output/step1_processed/intents_zh-hant.json", "name": "繁體中文"}
    ]
    
    print("\n" + "="*80)
    print("🌟 为每个意图创建独立知识库")
    print("="*80)
    print(f"当前配置:")
    print(f"  - 创建知识库: {'✅ 是' if CREATE_KB else '❌ 否（仅统计）'}")
    print(f"  - 处理数量限制: {MAX_INTENTS_TO_PROCESS if MAX_INTENTS_TO_PROCESS else '全部'}")
    print("="*80)
    
    print("\n请选择要处理的语言:")
    for idx, lang in enumerate(languages, 1):
        print(f"  {idx}. {lang['name']} ({lang['code']})")
    print(f"  4. 全部语言")
    
    choice = input("\n请输入选项 (1-4，默认为 1): ").strip() or "1"
    
    if choice == "4":
        # 处理所有语言
        for lang in languages:
            process_language_file(lang["file"], lang["code"])
            print("\n" + "="*80 + "\n")
    else:
        # 处理单个语言
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(languages):
                process_language_file(languages[idx]["file"], languages[idx]["code"])
            else:
                print("❌ 无效的选项")
        except ValueError:
            print("❌ 无效的输入")


if __name__ == "__main__":
    main()

